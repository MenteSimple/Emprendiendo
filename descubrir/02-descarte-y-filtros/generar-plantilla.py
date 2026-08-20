#!/usr/bin/env python3
"""Genera `plantilla-descarte.xlsx`, la hoja de trabajo del filtro.

El equipo la llena con lo que YA decidió —incluidos los motivos de cada
descarte— y la herramienta de IA agrega la columna «Revisión» sin tocar lo
demás. Sin motivos escritos no hay nada que auditar, así que el archivo
mismo hace cumplir la puerta de entrada.

    python3 generar-plantilla.py

Dos cosas de este archivo son defensas, no decoración:

1. Los rótulos de `G11:G13` en `Descartadas` son testigos: NO aparecen en el
   prompt, así que solo los puede citar quien abrió el archivo de verdad.
   Al cambiarlos hay que volver a correr el `grep` de cierre contra
   `prompt.txt` y actualizar la advertencia del README.
2. Los tres contadores de `H11:H13` son la firma del «nada más tocado».
   `H11` cuenta la columna que la herramienta sí escribe y tiene que subir;
   `H12` y `H13` cuentan trabajo del equipo y tienen que salir idénticos
   antes y después. Cualquier movimiento en esos dos delata una escritura
   fuera de `Revisión`.

La columna `¿Con quién hablamos?` lleva mensaje de entrada y NO lista
desplegable: el corte 2 existe para castigar las respuestas de catálogo, y
un desplegable es un catálogo. La única lista es la de `Origen`, donde el
juego de respuestas sí es cerrado y viene de la herramienta anterior.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SALIDA = Path(__file__).resolve().parent / "plantilla-descarte.xlsx"

AZUL, CREMA, BORDE_COLOR = "1F3A5F", "F1EEE7", "BFBFBF"
FILAS = 24
ULT = FILAS + 1  # última fila de datos

HOJAS = {
    "Descartadas": ["#", "La idea, en una línea", "Origen", "El motivo que dimos", "Revisión"],
    "Sobrevivientes": ["#", "La idea, en una línea", "Origen", "¿Con quién hablamos?", "¿Quién lo consigue?", "Revisión"],
    "Finalistas": ["Criterio", "① ", "② ", "③ ", "Revisión"],
}

ANCHOS = {
    "Descartadas": [5, 44, 9, 38, 42],
    "Sobrevivientes": [5, 40, 9, 30, 26, 38],
    "Finalistas": [34, 28, 28, 28, 40],
}

CRITERIOS = [
    "Deseabilidad — ¿alguien lo quiere de verdad?",
    "Factibilidad — ¿se puede construir con lo que existe hoy?",
    "Viabilidad — ¿hay forma de que se sostenga?",
    "Encaje con el equipo — ¿lo podemos hacer NOSOTROS?",
]

# Rótulos testigo: no deben aparecer nunca en el prompt, ni sueltos ni
# reconstruibles a partir de él. Al cambiarlos, correr:
#   grep -niE "tinta|puñ|huérfan" prompt.txt   → sin resultados
TESTIGOS = ["Casillas con tinta", "Puño del equipo", "Ideas huérfanas"]

MARCADOR = [
    f"=COUNTA(E2:E{ULT})",                          # Revisión: la única que sube
    f"=COUNTA(B2:D{ULT})",                          # idea + Origen + motivo: congelado
    f"=COUNTA(B2:B{ULT})-COUNTA(D2:D{ULT})",        # ideas sin motivo: congelado
]

AYUDA = {
    "Descartadas": (
        "Una fila por idea que botaron.\n\n"
        "• El motivo va con las palabras de ustedes, no en bonito. «Ya existe algo así» se escribe tal cual.\n"
        "• Sin motivo escrito la revisión no sirve: el filtro es el razonamiento, no la lista.\n"
        "• Origen viene de la herramienta anterior: P si la idea salió de un problema que ustedes observaron, S si salió de SCAMPER, X si la sugirió una IA.\n"
        "• La columna Revisión la llena la herramienta. Déjenla vacía.\n"
        "• El marcador de abajo cuenta tres cosas: las casillas de Revisión con texto, las celdas que escribieron ustedes, y las ideas que quedaron sin motivo. Anoten los dos últimos números antes de mandar el archivo: tienen que volver iguales."
    ),
    "Sobrevivientes": (
        "Las que pasaron los tres cortes.\n\n"
        "• «¿Con quién hablamos?» se responde con un nombre o un lugar concreto. Un tipo de persona no cuenta.\n"
        "• «¿Quién lo consigue?» es alguien del equipo, con nombre. Los dos van juntos: contacto sin responsable no es un contacto.\n"
        "• Origen: P si la idea salió de un problema que ustedes observaron, S si salió de SCAMPER, X si la sugirió una IA.\n"
        "• La columna Revisión la llena la herramienta."
    ),
    "Finalistas": (
        "Los tres que quedaron, comparados.\n\n"
        "• Escriban el nombre de cada finalista en la fila 1, al lado de su número.\n"
        "• Estos cuatro criterios no son binarios: tienen matices y de eso se trata la conversación.\n"
        "• El cuarto es el que más se deja en blanco y el que más pesa en un plazo corto.\n"
        "• «Nos parece interesante» no es encaje: encaje es qué sabe este equipo, a quién conoce y a qué tiene acceso.\n"
        "• El origen de cada finalista ya está en la hoja Sobrevivientes: acá no se repite."
    ),
}


def encabezar(hoja, titulos):
    for j, texto in enumerate(titulos, start=1):
        c = hoja.cell(row=1, column=j, value=texto)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 24
    hoja.freeze_panes = "A2"


def lista_origen(hoja, rango):
    """Desplegable P/S/X, en aviso y no en bloqueo, igual que en la herramienta 01."""
    dv = DataValidation(type="list", formula1='"P,S,X"', allow_blank=True, errorStyle="warning")
    dv.promptTitle = "Origen de la idea"
    dv.prompt = "P = problema que ustedes observaron · S = SCAMPER · X = se la sugirió una IA"
    dv.errorTitle = "Valor poco usual"
    dv.error = "Se esperaba P, S o X. Puede continuar, pero revise que no sea un error de dedo."
    hoja.add_data_validation(dv)
    dv.add(rango)


def aviso(hoja, rango, titulo, texto):
    """Mensaje de entrada sin lista: ayuda sin ofrecer la respuesta."""
    dv = DataValidation(showInputMessage=True, showErrorMessage=False, allow_blank=True)
    dv.promptTitle = titulo
    dv.prompt = texto
    hoja.add_data_validation(dv)
    dv.add(rango)


def main() -> None:
    wb = Workbook()
    borde = Border(*[Side(style="thin", color=BORDE_COLOR)] * 4)

    for nombre, titulos in HOJAS.items():
        hoja = wb.active if nombre == "Descartadas" else wb.create_sheet()
        hoja.title = nombre
        encabezar(hoja, titulos)

        filas = len(CRITERIOS) if nombre == "Finalistas" else FILAS
        for i in range(filas):
            r = i + 2
            if nombre == "Finalistas":
                hoja.cell(row=r, column=1, value=CRITERIOS[i]).alignment = Alignment(wrap_text=True, vertical="top")
            else:
                hoja.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center")
            for j in range(1, len(titulos) + 1):
                c = hoja.cell(row=r, column=j)
                c.border = borde
                if j > 1:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            hoja.row_dimensions[r].height = 30

        for j, w in enumerate(ANCHOS[nombre], start=1):
            hoja.column_dimensions[get_column_letter(j)].width = w

        # panel de ayuda, a la derecha y separado por una columna en blanco
        col = len(titulos) + 2
        hoja.column_dimensions[get_column_letter(col)].width = 52
        p = hoja.cell(row=2, column=col, value=AYUDA[nombre])
        p.alignment = Alignment(wrap_text=True, vertical="top")
        p.fill = PatternFill("solid", fgColor=CREMA)
        hoja.merge_cells(start_row=2, start_column=col, end_row=9, end_column=col)

    # testigos y marcador, solo en Descartadas: G11:G13 los rótulos, H11:H13 las fórmulas
    h = wb["Descartadas"]
    col_testigo = len(HOJAS["Descartadas"]) + 2
    for i, (rotulo, formula) in enumerate(zip(TESTIGOS, MARCADOR)):
        h.cell(row=11 + i, column=col_testigo, value=rotulo).font = Font(bold=True)
        h.cell(row=11 + i, column=col_testigo + 1, value=formula)
    h.column_dimensions[get_column_letter(col_testigo + 1)].width = 12

    lista_origen(wb["Descartadas"], f"C2:C{ULT}")
    lista_origen(wb["Sobrevivientes"], f"C2:C{ULT}")
    aviso(
        wb["Sobrevivientes"], f"D2:D{ULT}", "¿Con quién hablamos?",
        "Un nombre o un lugar concreto, escrito a mano. Un tipo de persona no cuenta.",
    )
    aviso(
        wb["Sobrevivientes"], f"E2:E{ULT}", "¿Quién lo consigue?",
        "Alguien del equipo, con nombre. Es la mitad que casi nadie escribe.",
    )

    wb.save(SALIDA)
    testigo_rango = f"{get_column_letter(col_testigo)}11:{get_column_letter(col_testigo)}13"
    print(f"✅ {SALIDA.name} · hojas: {', '.join(HOJAS)} · testigos en {testigo_rango} de Descartadas")


if __name__ == "__main__":
    main()
